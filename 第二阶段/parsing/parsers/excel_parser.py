"""XLSX 基础解析器；每行转换为制表符分隔文本。"""

from pathlib import Path

from 第二阶段.parsing.base import BaseParser, ParserDependencyError, ParserError
from 第二阶段.schemas.models import ParsedDocument


class ExcelParser(BaseParser):
    file_type = "xlsx"

    def parse(self, file_path: str | Path) -> ParsedDocument:
        path = self._validate_path(file_path)
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ParserDependencyError("XLSX 解析需要安装 openpyxl") from exc
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            blocks: list[str] = []
            for sheet in workbook.worksheets:
                blocks.append(f"[Sheet: {sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    if any(values):
                        blocks.append("\t".join(values))
            workbook.close()
        except Exception as exc:
            raise ParserError(f"XLSX 解析失败：{path}") from exc
        return self._build_document(path, "\n".join(blocks))

